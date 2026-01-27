from openpyxl import load_workbook
import sys

file_path = r'C:\Users\xuyaa\Desktop\BOCIASIV2.xlsx'
sheet_name = 'A'

try:
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    last_row = ws.max_row
    date_val = ws.cell(row=last_row, column=1).value
    
    print(f"Last row: {last_row}, Date: {date_val}")
    
    # Check if date is 2026-01-27 (or match string)
    if str(date_val).startswith('2026-01-27'):
        ws.delete_rows(last_row)
        wb.save(file_path)
        print("Successfully deleted the incomplete row for 2026-01-27")
    else:
        print("Last row is not today, skipping deletion.")
        
except Exception as e:
    print(f"Error: {e}")
