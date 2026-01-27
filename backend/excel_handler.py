"""
Excel操作模块 - 读写Excel文件，追加数据，备份
"""
import pandas as pd
import shutil
from datetime import datetime, timedelta
import os
import config

class ExcelHandler:
    """Excel文件处理类"""
    
    def __init__(self, excel_path=None, sheet_name=None):
        """
        初始化Excel处理器
        
        参数:
            excel_path: Excel文件路径，默认使用config中的路径
            sheet_name: 工作表名称，默认使用config中的名称
        """
        self.excel_path = excel_path or config.EXCEL_PATH
        self.sheet_name = sheet_name or config.SHEET_NAME
        self.df = None
    
    def read_excel(self):
        """读取Excel文件（只读取A-Q列的数据列）"""
        try:
            # 只读取前17列（A-Q列），不读取公式列（R-EW）
            self.df = pd.read_excel(
                self.excel_path, 
                sheet_name=self.sheet_name,
                usecols=range(17)  # 只读取列索引0-16（A-Q列）
            )
            return True
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {str(e)}")

    # ... (skipping some lines) ...

    def get_last_date(self):
        """
        获取Excel中的最后一个日期
        
        返回:
            str: 日期字符串，格式 "YYYY-MM-DD"
        """
        if self.df is None:
            self.read_excel()
        
        if len(self.df) == 0:
            return None
        
        # 获取第一列（日期列）的最后一个值
        last_date = self.df.iloc[-1, 0]
        
        # 转换为字符串格式
        if isinstance(last_date, datetime):
            return last_date.strftime('%Y-%m-%d')
        elif isinstance(last_date, pd.Timestamp):
            return last_date.strftime('%Y-%m-%d')
        else:
            return str(last_date)
    
    def get_next_row_number(self):
        """
        获取下一行的行号（Excel格式，从1开始）
        
        返回:
            int: 下一行行号
        """
        if self.df is None:
            self.read_excel()
        
        # +2 是因为: +1 for header, +1 for next row
        return len(self.df) + 2
    
    def append_data(self, data_dict):
        """
        追加一行数据到DataFrame
        
        参数:
            data_dict: 数据字典，keys应与config.COLUMN_MAPPING的keys匹配
        """
        if self.df is None:
            self.read_excel()
        
        # 构建新行数据（按照Excel列顺序）
        # self.df 是读取的前17列（A-Q），所以这里必须初始化为17个元素的列表
        new_row = [None] * 17
        
        for key, col_idx in config.COLUMN_MAPPING.items():
            if key in data_dict:
                # 确保索引在范围内
                if col_idx < 17:
                    new_row[col_idx] = data_dict[key]
        
        # 转换为DataFrame并追加
        new_df = pd.DataFrame([new_row], columns=self.df.columns)
        self.df = pd.concat([self.df, new_df], ignore_index=True)
    
    def save_excel(self):
        """保存Excel文件，使用openpyxl直接追加数据行并复制公式"""
        if self.df is None:
            raise Exception("没有数据可保存")
        
        try:
            from openpyxl import load_workbook
            
            # 打开现有的Excel文件
            wb = load_workbook(self.excel_path)
            ws = wb[self.sheet_name]
            
            # 获取当前最后一行
            current_last_row = ws.max_row
            
            # 计算需要追加的新行数
            # df的行数 - (Excel行数 - 标题行)
            excel_data_rows = current_last_row - 1  # 减去标题行
            df_data_rows = len(self.df)
            new_rows_count = df_data_rows - excel_data_rows
            
            if new_rows_count <= 0:
                print("   没有新数据需要追加")
                wb.close()
                return True
            
            print(f"   需要追加 {new_rows_count} 行数据")
            
            # 追加新行
            for i in range(new_rows_count):
                row_idx_in_df = excel_data_rows + i  # df中的行索引
                new_row_in_excel = current_last_row + i + 1  # Excel中的新行号
                
                print(f"   追加第{new_row_in_excel}行...")
                
                # 1. 写入A-Q列的数据（17列）
                for col_idx in range(17):
                    cell_value = self.df.iloc[row_idx_in_df, col_idx]
                    cell = ws.cell(row=new_row_in_excel, column=col_idx + 1)
                    
                    # N列（MA20，索引13）：设置为百分比格式（不除以100，直接显示）
                    if col_idx == 13:  # N列：MA20宽度
                        if cell_value is not None and not pd.isna(cell_value):
                            cell.value = cell_value
                            cell.number_format = '0.00"%"'  # 自定义格式：数字后面加%号
                        else:
                            cell.value = cell_value
                    else:
                        cell.value = cell_value
                
                # 1.1 特殊处理：G列融资余额（第7列，索引6）
                # 如果当日融资余额为空，使用上一日的数据
                if new_row_in_excel > 2:  # 不是第一行数据
                    margin_col = 7  # G列
                    current_margin = ws.cell(row=new_row_in_excel, column=margin_col).value
                    if current_margin is None or (isinstance(current_margin, float) and pd.isna(current_margin)):
                        # 使用上一行的融资余额
                        prev_margin = ws.cell(row=new_row_in_excel - 1, column=margin_col).value
                        ws.cell(row=new_row_in_excel, column=margin_col, value=prev_margin)
                        print(f"      G列融资余额缺失，使用上一日数据: {prev_margin}")
                
                # 1.2 特殊处理：DN列序号（第118列）
                # DN列 = 4*26 + 14 = 118
                dn_col = 118  # DN列
                if new_row_in_excel == 4637:
                    # 第4637行填入3171
                    ws.cell(row=new_row_in_excel, column=dn_col, value=3171)
                elif new_row_in_excel > 2:
                    # 其他行：上一行的值 + 1
                    prev_dn = ws.cell(row=new_row_in_excel - 1, column=dn_col).value
                    if prev_dn and isinstance(prev_dn, (int, float)):
                        ws.cell(row=new_row_in_excel, column=dn_col, value=int(prev_dn) + 1)
                    else:
                        # 如果上一行没有值，根据行号计算
                        # 4637行=3171，所以公式是：3171 + (当前行 - 4637)
                        dn_value = 3171 + (new_row_in_excel - 4637)
                        ws.cell(row=new_row_in_excel, column=dn_col, value=dn_value)
                
                # 2. 从上一行复制Q-EW列的公式（如果上一行存在）
                if new_row_in_excel > 2:  # 确保不是第一行数据
                    source_row = new_row_in_excel - 1
                    target_row = new_row_in_excel
                    
                    # Q列=17到EW列=157
                    formula_start_col = 17
                    formula_end_col = 157
                    
                    import re
                    
                    for col in range(formula_start_col, formula_end_col + 1):
                        source_cell = ws.cell(row=source_row, column=col)
                        target_cell = ws.cell(row=target_row, column=col)
                        
                        # 如果源单元格有公式，复制并更新行号
                        if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                            formula = source_cell.value
                            
                            # 替换公式中的行号引用
                            # 匹配模式如：A4637, $A4637, A$4637, $A$4637
                            def replace_row_number(match):
                                col_ref = match.group(1)  # 列引用（如A, $A等）
                                old_row = int(match.group(2))  # 旧行号
                                
                                # 检查行号前是否有$符号（绝对引用）
                                # 如果列引用以$结尾或者整个引用是绝对的，行号也应该是绝对的
                                # 例如：$A$4637 中，列和行都是绝对的
                                # A$4637 中，只有行是绝对的
                                
                                # 检查原公式中这个位置的行号前是否有$
                                full_match = match.group(0)  # 完整匹配，如 $A$4637
                                
                                # 如果是绝对行引用（行号前有$），保持行号不变
                                if f"${old_row}" in full_match:
                                    return full_match  # 保持原样
                                
                                # 相对引用：更新行号
                                row_diff = target_row - source_row
                                new_row_num = old_row + row_diff
                                
                                return f"{col_ref}{new_row_num}"
                            
                            # 正则替换：匹配列字母+行号
                            # 模式：(\$?[A-Z]+\$?)(\d+)
                            updated_formula = re.sub(r'(\$?[A-Z]+\$?)(\d+)', replace_row_number, formula)
                            
                            target_cell.value = updated_formula
                            
                            # 复制格式
                            if source_cell.has_style:
                                target_cell.font = source_cell.font.copy()
                                target_cell.border = source_cell.border.copy()
                                target_cell.fill = source_cell.fill.copy()
                                target_cell.number_format = source_cell.number_format
                                target_cell.protection = source_cell.protection.copy()
                                target_cell.alignment = source_cell.alignment.copy()
            
            # 保存工作簿
            wb.save(self.excel_path)
            wb.close()
            
            print(f"   ✅ Excel文件已保存（追加了{new_rows_count}行，包含公式）")
            return True
        except Exception as e:
            raise Exception(f"保存Excel文件失败: {str(e)}")

    def update_margin_for_date(self, date_str, new_value):
        """
        更新指定日期的融资余额
        
        参数:
            date_str: 日期字符串
            new_value: 新的融资余额值
        """
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_path)
            ws = wb[self.sheet_name]
            
            # 从最后一行往上找
            max_row = ws.max_row
            target_row = None
            
            # 只检查最后20行
            for r in range(max_row, max(1, max_row-20), -1):
                cell_date = ws.cell(row=r, column=1).value
                # 处理日期格式
                d_str = ""
                if hasattr(cell_date, 'strftime'):
                    d_str = cell_date.strftime('%Y-%m-%d')
                else:
                    d_str = str(cell_date).split(' ')[0]
                
                if d_str == date_str:
                    target_row = r
                    break
            
            if target_row:
                 # 融资余额在第7列 (G列)
                 ws.cell(row=target_row, column=7, value=new_value)
                 wb.save(self.excel_path)
                 print(f"   ✅ 已修正 {date_str} 的融资余额为: {new_value}")
                 return True
            else:
                print(f"   ⚠️ 未找到日期 {date_str}，无法更新融资余额")
                return False
                
        except Exception as e:
            print(f"   ❌ 更新融资余额失败: {str(e)}")
            return False
    
    def backup_excel(self):
        """
        备份Excel文件
        
        返回:
            str: 备份文件路径
        """
        backup_path = config.get_backup_filename()
        
        try:
            shutil.copy2(self.excel_path, backup_path)
            print(f"✅ 备份文件已创建: {backup_path}")
            
            # 清理旧备份
            self._cleanup_old_backups()
            
            return backup_path
        except Exception as e:
            raise Exception(f"备份文件失败: {str(e)}")
    
    def _cleanup_old_backups(self):
        """清理超过保留期限的备份文件"""
        try:
            cutoff_date = datetime.now() - timedelta(days=config.BACKUP_RETENTION_DAYS)
            
            for filename in os.listdir(config.BACKUP_DIR):
                if filename.startswith('BOCIASIV2_backup_') and filename.endswith('.xlsx'):
                    file_path = os.path.join(config.BACKUP_DIR, filename)
                    file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
                    
                    if file_mtime < cutoff_date:
                        os.remove(file_path)
                        print(f"  清理旧备份: {filename}")
        except Exception as e:
            print(f"  ⚠️ 清理备份文件时出错: {str(e)}")
    
    def validate_data(self, data_dict):
        """
        验证数据的完整性
        
        参数:
            data_dict: 数据字典
        
        返回:
            tuple: (is_valid, missing_fields, message)
        """
        required_fields = ['date', 'close', 'turnover']
        missing_fields = []
        
        for field in required_fields:
            if field not in data_dict or data_dict[field] is None:
                missing_fields.append(field)
        
        if missing_fields:
            return False, missing_fields, f"缺少必要字段: {', '.join(missing_fields)}"
        
        return True, [], "数据验证通过"
    
    def print_summary(self):
        """打印数据摘要"""
        if self.df is None:
            self.read_excel()
        
        print(f"\n📊 Excel 数据摘要:")
        print(f"  文件路径: {self.excel_path}")
        print(f"  工作表: {self.sheet_name}")
        print(f"  总行数: {len(self.df) + 1} (包含标题行)")
        print(f"  数据行数: {len(self.df)}")
        print(f"  列数: {len(self.df.columns)}")
        
        if len(self.df) > 0:
            last_date = self.get_last_date()
            print(f"  最后日期: {last_date}")
            print(f"  下一行行号: {self.get_next_row_number()}")


# 便捷函数
def update_excel_with_data(data_list, backup=True):
    """
    更新Excel文件（追加多行数据）
    
    参数:
        data_list: 数据字典列表
        backup: 是否在更新前备份
    
    返回:
        bool: 是否成功
    """
    handler = ExcelHandler()
    
    try:
        # 读取Excel
        handler.read_excel()
        
        # 备份
        if backup:
            handler.backup_excel()
        
        # 追加数据
        for data in data_list:
            handler.append_data(data)
        
        # 保存
        handler.save_excel()
        
        return True
    except Exception as e:
        print(f"❌ 更新Excel失败: {str(e)}")
        return False
