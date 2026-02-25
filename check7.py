"""
 快速验证：检查 Excel 文件中 AF 列的前几个单元格是公式还是数值
 使用 openpyxl READ_ONLY=False 以访问公式内容
"""
import openpyxl

excel_path = r'C:\Users\xuyaa\Desktop\BOCIASIV2.xlsx'

print("Loading with data_only=False (shows formulas)...")
wb = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)
ws = wb['A']

print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# 检查几个关键单元格
# Excel row 2195 (pandas 2194) 是第一个数据行附近
test_rows = [2194, 2195, 2196, 4654, 4655]
test_cols = {
    'A': 1, 'C': 3, 'AF': 32, 'DC': 107, 'DD': 108
}

for r in test_rows:
    row_vals = []
    for col_name, col_idx in test_cols.items():
        cell = ws.cell(row=r, column=col_idx)
        row_vals.append(f"{col_name}={str(cell.value)[:30]!r}")
    print(f"Row{r}: " + " | ".join(row_vals))

wb.close()
print("Done.")
