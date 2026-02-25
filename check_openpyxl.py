"""
直接用 openpyxl 读取特定单元格，确认 Excel 中的实际值
"""
import openpyxl

excel_path = r'C:\Users\xuyaa\Desktop\BOCIASIV2.xlsx'
print("Loading workbook (data_only=True)...")
wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
ws = wb['A']

# Excel 行号从1开始，第2194行是数据起始行
# 检查行2194到2200，以及最后几行
# 列映射: A=1, C=3, AF=32, DC=107, DD=108

col_map = {
    'A(date)': 1,
    'C(close)': 3,
    'AF(equity)': 32,
    'BN(margin)': 66,
    'CB(turnover)': 80,
    'CP(ma20)': 94,
    'DC(fast)': 107,
    'DD(slow)': 108,
}

results = []
results.append("=== openpyxl data_only read ===")
results.append(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
results.append("")

# 检查2192~2202行
results.append("Rows 2192~2205:")
for excel_row in range(2192, 2206):
    vals = []
    for col_name, col_idx in col_map.items():
        cell = ws.cell(row=excel_row, column=col_idx)
        vals.append(f"{col_name}={cell.value!r:.20}")
    results.append(f"  ExcelRow{excel_row}: " + " | ".join(vals))

results.append("")
results.append("Last 5 rows:")
max_row = ws.max_row
for excel_row in range(max(1, max_row - 4), max_row + 1):
    vals = []
    for col_name, col_idx in col_map.items():
        cell = ws.cell(row=excel_row, column=col_idx)
        vals.append(f"{col_name}={str(cell.value)[:15]!r}")
    results.append(f"  ExcelRow{excel_row}: " + " | ".join(vals))

wb.close()

output = '\n'.join(results)
with open('openpyxl_result.txt', 'w', encoding='utf-8') as f:
    f.write(output)

# Print line by line
for r in results:
    print(r)
