"""
把 BOCIASIV2.xlsx 中 2026-02-12 及之后的行全部删除，
然后重新从 Wind 拉取数据更新。
"""
import sys
import os

# 加入 backend 目录到 path
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backend'))

from openpyxl import load_workbook
import config

EXCEL_PATH = config.EXCEL_PATH
SHEET_NAME = config.SHEET_NAME
ROLLBACK_DATE = '2026-02-12'   # 从这天（含）开始删除

print(f"Excel: {EXCEL_PATH}")
print(f"将删除 {ROLLBACK_DATE} 及之后的所有行...")

wb = load_workbook(EXCEL_PATH)
ws = wb[SHEET_NAME]

max_row = ws.max_row
rows_to_delete = []

for r in range(2, max_row + 1):      # 跳过第1行（标题行）
    cell_date = ws.cell(row=r, column=1).value
    if cell_date is None:
        continue
    # 转成字符串
    if hasattr(cell_date, 'strftime'):
        d_str = cell_date.strftime('%Y-%m-%d')
    else:
        d_str = str(cell_date)[:10]
    
    if d_str >= ROLLBACK_DATE:
        rows_to_delete.append(r)

print(f"找到 {len(rows_to_delete)} 行需要删除（从Excel行 {rows_to_delete[0] if rows_to_delete else 'N/A'}）")

if not rows_to_delete:
    print("没有需要删除的行，退出。")
    wb.close()
    sys.exit(0)

# 从后往前删，避免行号错位
for r in reversed(rows_to_delete):
    ws.delete_rows(r)

print(f"删除完成，当前总行数: {ws.max_row}")

# 确认最后一行
last_date_cell = ws.cell(row=ws.max_row, column=1).value
print(f"删除后最后一行日期: {last_date_cell}")

wb.save(EXCEL_PATH)
wb.close()
print("已保存！")
